import argparse
import csv
import json
from collections import Counter
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("openpyxl is required to run this scaffold.") from exc


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_CONFIG = ROOT / "config" / "pilot_config.json"


def load_config(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def read_sheet_as_records(wb, sheet_name: str, header_row: int = 4) -> list[dict]:
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[header_row]]
    records = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(value not in (None, "") for value in row):
            continue
        record = {}
        for idx, header in enumerate(headers):
            if header is None:
                continue
            record[str(header)] = row[idx] if idx < len(row) else None
        records.append(record)
    return records


def normalize(value):
    if value is None:
        return ""
    return str(value).strip()


def write_csv(path: Path, rows: list[dict], fieldnames: list[str]):
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def check_required_sheets(wb, config):
    issues = []
    present = set(wb.sheetnames)
    for sheet in config["required_sheets"]:
        if sheet not in present:
            issues.append({
                "severity": "fail",
                "area": "workbook_structure",
                "issue": f"Missing required sheet: {sheet}",
                "action": "Restore the pilot workbook template or add the missing sheet."
            })
    return issues


def check_candidate_indicators(records):
    retained = [r for r in records if normalize(r.get("candidate_decision")).lower() == "retain"]
    pending = [r for r in records if normalize(r.get("candidate_decision")).lower() == "pending"]
    issues = []
    if not retained:
        issues.append({
            "severity": "warning" if pending else "fail",
            "area": "candidate_indicators",
            "issue": "No retained dental/oral indicator is recorded.",
            "action": "Retain one simple prefecture-year dental/oral indicator before continuing."
        })
    return retained, issues


def check_rule_verification(records):
    verified = []
    issues = []
    for r in records:
        status = normalize(r.get("rule_status")).lower()
        eligible = normalize(r.get("primary_bounds_eligible")).lower()
        threshold = r.get("threshold_T")
        symbol = normalize(r.get("suppression_symbol"))
        if status == "verified" and eligible == "yes":
            if threshold in (None, "") or not symbol:
                issues.append({
                    "severity": "fail",
                    "area": "rule_verification",
                    "issue": f"Verified rule {r.get('rule_key')} lacks threshold_T or suppression_symbol.",
                    "action": "Fill threshold_T and suppression_symbol before primary bounds."
                })
            else:
                verified.append(r)
    if not verified:
        issues.append({
            "severity": "fail",
            "area": "rule_verification",
            "issue": "No verified and primary-eligible rule is available.",
            "action": "Do not run primary known-censoring bounds until at least one rule is verified."
        })
    return verified, issues


def check_cell_states(records, allowed_states):
    counts = Counter()
    issues = []
    for r in records:
        state = normalize(r.get("cell_state")).lower()
        if not state:
            continue
        counts[state] += 1
        if state not in allowed_states:
            issues.append({
                "severity": "fail",
                "area": "cell_state_trial",
                "issue": f"Unknown cell_state '{state}' for cell_id {r.get('cell_id')}.",
                "action": "Use only controlled cell-state categories."
            })
    if counts["observed"] == 0:
        issues.append({
            "severity": "warning",
            "area": "cell_state_trial",
            "issue": "No observed cells are present in the trial.",
            "action": "Include at least one observed cell for comparison."
        })
    if counts["suppressed"] == 0:
        issues.append({
            "severity": "warning",
            "area": "cell_state_trial",
            "issue": "No suppressed cells are present in the trial.",
            "action": "Include at least one suppressed cell or choose another indicator/table."
        })
    return counts, issues


def check_bounds(records):
    rows = []
    issues = []
    suppressed_with_bounds = 0
    for r in records:
        cell_id = r.get("cell_id")
        state = normalize(r.get("cell_state")).lower()
        lower = r.get("count_lower")
        upper = r.get("count_upper")
        threshold = r.get("threshold_T")
        rule = normalize(r.get("lower_bound_rule")).lower()
        status = "not_checked"
        detail = ""

        if state == "observed":
            observed = r.get("observed_count")
            if observed in (None, ""):
                status = "pending"
                detail = "Observed cell lacks observed_count."
            elif lower not in (None, "") or upper not in (None, ""):
                status = "check"
                detail = "Observed cell should usually use lower=upper=observed_count."
            else:
                status = "pending"
                detail = "Fill count_lower=count_upper=observed_count."
        elif state == "suppressed":
            if threshold in (None, ""):
                status = "fail"
                detail = "Suppressed cell lacks threshold_T."
            elif rule not in {"zero_possible", "event_exists"}:
                status = "fail"
                detail = "Suppressed cell must use zero_possible or event_exists."
            elif lower in (None, "") or upper in (None, ""):
                status = "fail"
                detail = "Suppressed cell lacks count bounds."
            else:
                try:
                    t = float(threshold)
                    lo = float(lower)
                    hi = float(upper)
                    expected_lo = 0 if rule == "zero_possible" else 1
                    expected_hi = t - 1
                    if lo == expected_lo and hi == expected_hi:
                        status = "pass"
                        detail = "Suppressed-cell interval matches rule."
                        suppressed_with_bounds += 1
                    else:
                        status = "fail"
                        detail = f"Expected [{expected_lo}, {expected_hi}] but found [{lo}, {hi}]."
                except Exception:
                    status = "fail"
                    detail = "Bounds or threshold are not numeric."
        else:
            if lower not in (None, "") or upper not in (None, ""):
                status = "fail"
                detail = "Non-primary state has count bounds assigned."
            else:
                status = "pass"
                detail = "No primary bounds assigned."

        rows.append({
            "cell_id": cell_id,
            "cell_state": state,
            "lower_bound_rule": rule,
            "threshold_T": threshold,
            "count_lower": lower,
            "count_upper": upper,
            "check_status": status,
            "detail": detail
        })
        if status == "fail":
            issues.append({
                "severity": "fail",
                "area": "bounds_trial",
                "issue": f"Bounds check failed for {cell_id}: {detail}",
                "action": "Correct Bounds_Trial before expanding pilot."
            })
    return rows, suppressed_with_bounds, issues


def check_naive(records, required):
    strategies = {normalize(r.get("strategy")).lower() for r in records}
    required_normalized = [s.lower() for s in required]
    missing = [s for s in required_normalized if s not in strategies]
    issues = []
    if missing:
        issues.append({
            "severity": "fail",
            "area": "naive_comparison",
            "issue": "Missing naive strategies: " + ", ".join(missing),
            "action": "Add all four primary naive strategies."
        })
    ready = [
        r for r in records
        if normalize(r.get("ready_for_full_run")).lower() == "yes"
    ]
    return strategies, ready, issues


def recommend(summary, issues, config):
    fail_areas = {i["area"] for i in issues if i["severity"] == "fail"}
    warning_areas = {i["area"] for i in issues if i["severity"] == "warning"}

    if "rule_verification" in fail_areas:
        return "HOLD_PRIMARY_BOUNDS", "Rule verification is not strong enough for primary known-censoring bounds."
    if "bounds_trial" in fail_areas:
        return "REVISE_INDICATOR", "Bounds logic failed for the current pilot indicator or entries."
    if "candidate_indicators" in fail_areas:
        return "REVISE_INDICATOR", "No retained candidate indicator is available."
    if "naive_comparison" in fail_areas:
        return "REVISE_INDICATOR", "Naive comparison setup is incomplete."
    if summary["suppressed_cells_with_bounds"] >= config["go_criteria"]["min_suppressed_cells_with_bounds"]:
        if warning_areas:
            return "GO_WITH_MINOR_WARNINGS", "Pilot can expand, but warnings should be documented."
        return "GO", "Pilot supports expansion to full dental/oral extraction."
    return "STOP_REDESIGN", "No suppressed cell with valid bounds was demonstrated."


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True, help="Path to populated Dental_Oral_Pilot_Run_Workbook.xlsx")
    parser.add_argument("--outdir", default=str(ROOT / "results"), help="Output directory")
    parser.add_argument("--config", default=str(DEFAULT_CONFIG), help="Config JSON path")
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)
    config = load_config(Path(args.config))

    wb = load_workbook(workbook_path, data_only=True)
    issues = []
    issues.extend(check_required_sheets(wb, config))

    if issues:
        summary = {"workbook": str(workbook_path), "fatal_structure_issue": True}
    else:
        candidates = read_sheet_as_records(wb, "Candidate_Indicators")
        rules = read_sheet_as_records(wb, "Rule_Verification")
        cells = read_sheet_as_records(wb, "Cell_State_Trial")
        bounds = read_sheet_as_records(wb, "Bounds_Trial")
        naive = read_sheet_as_records(wb, "Naive_Comparison")

        retained, new_issues = check_candidate_indicators(candidates)
        issues.extend(new_issues)
        verified_rules, new_issues = check_rule_verification(rules)
        issues.extend(new_issues)
        state_counts, new_issues = check_cell_states(cells, set(config["allowed_cell_states"]))
        issues.extend(new_issues)
        bounds_rows, suppressed_with_bounds, new_issues = check_bounds(bounds)
        issues.extend(new_issues)
        strategies, naive_ready, new_issues = check_naive(naive, config["required_naive_strategies"])
        issues.extend(new_issues)

        write_csv(outdir / "cell_state_counts.csv",
                  [{"cell_state": k, "count": v} for k, v in sorted(state_counts.items())],
                  ["cell_state", "count"])
        write_csv(outdir / "bounds_trial_checks.csv", bounds_rows,
                  ["cell_id", "cell_state", "lower_bound_rule", "threshold_T", "count_lower", "count_upper", "check_status", "detail"])

        summary = {
            "workbook": str(workbook_path),
            "retained_indicators": len(retained),
            "verified_primary_rules": len(verified_rules),
            "cell_state_counts": dict(state_counts),
            "suppressed_cells_with_bounds": suppressed_with_bounds,
            "naive_strategies_present": sorted(strategies),
            "naive_strategies_ready_for_full_run": len(naive_ready),
            "issue_count": len(issues)
        }

    recommendation, reason = recommend(summary, issues, config)
    summary["recommendation"] = recommendation
    summary["recommendation_reason"] = reason

    with (outdir / "pilot_summary.json").open("w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2, ensure_ascii=False)

    write_csv(outdir / "pilot_issues.csv", issues,
              ["severity", "area", "issue", "action"])

    with (outdir / "go_nogo_recommendation.txt").open("w", encoding="utf-8") as f:
        f.write(f"{recommendation}\n\n{reason}\n")

    print(json.dumps(summary, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
